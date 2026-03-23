#!/usr/bin/env python3
"""
TerraFusion OS — CostForge Schedule Seeder (Benton County)
===========================================================
Reads Benton County's custom cost approach Excel files and seeds TerraFusion
Supabase CostForge tables. These are Benton's own schedules — built without
Marshall & Swift — that form the foundation of TerraFusion CostForge.

Source files (all in 'Schedules and Matrices' folder):
  Cost Approach.xlsx              → costforge_residential_schedules
  Section 11-18, 61, 64 xlsx     → costforge_commercial_schedules
  Depreciation.xlsx               → costforge_depreciation
  local and current cost multipliers.xlsx → costforge_cost_multipliers
  Refinement Matrices.xlsx        → costforge_refinements (HVAC, sprinklers, elevators)
  PACS detail type codes-8-24.xlsx → costforge_imprv_type_codes

Requirements:
  py -3.12 -m pip install requests openpyxl python-dotenv

Usage:
  py -3.12 scripts/seed_costforge_benton.py
  py -3.12 scripts/seed_costforge_benton.py --dry-run
  py -3.12 scripts/seed_costforge_benton.py --table residential
  py -3.12 scripts/seed_costforge_benton.py --table commercial
  py -3.12 scripts/seed_costforge_benton.py --table depreciation
  py -3.12 scripts/seed_costforge_benton.py --table multipliers
  py -3.12 scripts/seed_costforge_benton.py --table refinements
  py -3.12 scripts/seed_costforge_benton.py --table type_codes
"""

from __future__ import annotations

import os
import sys
import time
import argparse
from pathlib import Path
from typing import Any

# ── Dependency check ──────────────────────────────────────────────────────────

MISSING: list[str] = []
try:
    import requests
except ImportError:
    MISSING.append("requests")
try:
    import openpyxl
except ImportError:
    MISSING.append("openpyxl")
if MISSING:
    print(f"[seed_costforge] Missing: {', '.join(MISSING)}")
    print("  Install: py -3.12 -m pip install requests openpyxl python-dotenv")
    sys.exit(1)

try:
    from dotenv import load_dotenv
    _seed_env = Path(__file__).parent / ".env.seed"
    if _seed_env.exists():
        load_dotenv(_seed_env)
    load_dotenv()
except ImportError:
    _seed_env = Path(__file__).parent / ".env.seed"
    if _seed_env.exists():
        for _line in _seed_env.read_text().splitlines():
            _line = _line.strip()
            if _line and not _line.startswith("#") and "=" in _line:
                _k, _, _v = _line.partition("=")
                os.environ.setdefault(_k.strip(), _v.strip())

# ── Configuration ─────────────────────────────────────────────────────────────

SUPABASE_URL = os.getenv("SUPABASE_URL", "https://udjoodlluygvlqccwade.supabase.co").rstrip("/")
SERVICE_KEY  = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")

SCHEDULES_DIR = Path(os.getenv(
    "COSTFORGE_SCHEDULES_DIR",
    r"e:\Files of SQL\Files of SQL\Schedules and Matrices"
))

BENTON_COUNTY_ID = os.getenv("BENTON_COUNTY_ID", "842a6c54-c7c0-4b2d-aa43-0e3ba63fa57d")
BATCH_SIZE = 250

# Commercial section files: section_id → filename
COMMERCIAL_SECTIONS: dict[int, str] = {
    11: "Section 11.xlsx",
    12: "Section 12.xlsx",
    13: "Section 13.xlsx",
    14: "Section 14.xlsx",
    15: "Section 15.xlsx",
    16: "Section 16.xlsx",
    17: "Section 17.xlsx",
    18: "Section 18.xlsx",
    61: "Section 61 Tanks.xlsx",
    64: "Section 64 Service Stations Carwashes.xlsx",
}

# Construction class labels as they appear in the Excel files
CONSTRUCTION_CLASSES = [
    ("Class A",  "A"),  # Structural Steel Frame
    ("Class B",  "B"),  # Reinforced Concrete
    ("Class C",  "C"),  # Masonry
    ("Class D",  "D"),  # Wood or Steel Frame
    ("Class S",  "S"),  # Pre-Engineered Steel
    ("Class P",  "P"),  # Pole Frame
]

# Residential exterior wall types (header row in Cost Approach.xlsx)
RESIDENTIAL_EXT_WALLS = [
    "Plywood or Hardboard",
    "Metal or Vinyl Siding",
    "Stucco",
    "Wood Siding",
    "Wood Shingles",
    "Synth. Plaster",
]

# ── HTTP helpers ──────────────────────────────────────────────────────────────

from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter

_session = requests.Session()
_retry = Retry(total=5, backoff_factor=1.0, status_forcelist=[429, 502, 503, 504],
               allowed_methods=["POST", "GET"])
_session.mount("https://", HTTPAdapter(max_retries=_retry))


def _headers() -> dict[str, str]:
    return {
        "apikey": SERVICE_KEY,
        "Authorization": f"Bearer {SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal,resolution=merge-duplicates",
    }


def _upsert_batch(table: str, rows: list, on_conflict: str) -> None:
    if not rows:
        return
    url = f"{SUPABASE_URL}/rest/v1/{table}?on_conflict={on_conflict}"
    resp = _session.post(url, headers=_headers(), json=rows, timeout=120)
    if resp.status_code not in (200, 201):
        snippet = resp.text[:400].replace(SERVICE_KEY, "***")
        raise RuntimeError(f"upsert {table} [{resp.status_code}]: {snippet}")


def bulk_upsert(table: str, rows: list, on_conflict: str,
                label: str = "", dry_run: bool = False) -> None:
    if dry_run:
        print(f"  DRY-RUN {label or table}: {len(rows):,} rows")
        return
    total = len(rows)
    t0 = time.time()
    for i in range(0, total, BATCH_SIZE):
        _upsert_batch(table, rows[i: i + BATCH_SIZE], on_conflict)
        sent = min(i + BATCH_SIZE, total)
        print(f"  {label or table}: {sent:,}/{total:,} — {time.time()-t0:.1f}s", end="\r")
    print(f"  {label or table}: {total:,} rows written in {time.time()-t0:.1f}s    ")


def _safe_float(val: Any) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _safe_int(val: Any) -> int | None:
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _clean_str(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip().rstrip().replace("\n", " ")
    return s if s else None

# ── 1. Residential schedules ──────────────────────────────────────────────────

def seed_residential(dry_run: bool = False) -> None:
    """
    Parse Cost Approach.xlsx.
    Sheet name = quality_grade.
    Row[0] col[0] = 'Total area', cols[1..] = ext_wall_types.
    Subsequent rows: col[0] = min_area (sqft), cols[1..] = unit_cost.
    """
    path = SCHEDULES_DIR / "Cost Approach.xlsx"
    print(f"\n[1/6] costforge_residential_schedules ({path.name})")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows_out: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        # Find header row (contains 'Total area')
        header_row = None
        header_idx  = 0
        for i, row in enumerate(all_rows[:5]):
            if row and any("Total area" in str(c) for c in row if c is not None):
                header_row = row
                header_idx = i
                break
        if header_row is None:
            continue

        # Map column index → wall type (skip col 0 = area column)
        wall_types: dict[int, str] = {}
        for col_i, cell in enumerate(header_row):
            if col_i == 0:
                continue
            wall = _clean_str(cell)
            if wall:
                wall_types[col_i] = wall

        # Data rows
        quality_grade = _clean_str(sheet_name)
        for row in all_rows[header_idx + 1:]:
            if not row:
                continue
            min_area = _safe_int(row[0])
            if min_area is None:
                continue
            for col_i, wall in wall_types.items():
                if col_i >= len(row):
                    continue
                unit_cost = _safe_float(row[col_i])
                if unit_cost is None:
                    continue
                rows_out.append({
                    "county_id":     BENTON_COUNTY_ID,
                    "quality_grade": quality_grade,
                    "min_area":      min_area,
                    "ext_wall_type": wall,
                    "unit_cost":     unit_cost,
                    "source_file":   path.name,
                })

    wb.close()
    bulk_upsert("costforge_residential_schedules", rows_out,
                "county_id,quality_grade,min_area,ext_wall_type", dry_run=dry_run)


# ── 2. Commercial schedules (Sections 11-18, 61, 64) ─────────────────────────

def _parse_commercial_sheet(
    ws: "openpyxl.worksheet.worksheet.Worksheet",
    section_id: int,
    sheet_name: str,
    file_name: str,
    county_id: str,
) -> list[dict]:
    """
    Parse one occupancy sheet from a Section xlsx.
    Row structure:
      Row 1: occupancy name (e.g. 'Homes for the Elderly (330) pg 13')
      Row 2: PACS occupancy code (e.g. 'HOMEELDY')
      Row 3+: metadata notes
      Then a blank row, then the data table:
        Col 0 = construction class label
        Col 1 = Low unit cost
        Col 2 = % diff to Avg
        Col 3 = Avg unit cost
        Col 4 = % diff to Good
        Col 5 = Good unit cost
        Col 6 = % diff to Exc
        Col 7 = Exc unit cost
    Occupancy code is the sheet name (e.g. '330').
    """
    rows_out: list[dict] = []
    all_rows = list(ws.iter_rows(max_row=50, values_only=True))

    if not all_rows:
        return rows_out

    # Extract occupancy description from row 0
    occupancy_desc = _clean_str(all_rows[0][0]) if all_rows[0] else None
    occupancy_code = sheet_name  # sheet name is the occupancy number

    # Find the data grid: look for a row containing a construction class label
    class_prefixes = [c[0] for c in CONSTRUCTION_CLASSES]
    data_start = None
    for i, row in enumerate(all_rows):
        if row and row[0] is not None:
            cell_str = str(row[0]).strip()
            if any(cell_str.startswith(pref) for pref in class_prefixes):
                data_start = i
                break

    if data_start is None:
        return rows_out

    # Quality column mapping: Low=1, Avg=3, Good=5, Exc=7
    quality_map = {1: "Low", 3: "Average", 5: "Good", 7: "Excellent"}
    pct_diff_map = {1: 2, 3: 4, 5: 6}  # col idx of % diff immediately after each quality

    for row in all_rows[data_start:]:
        if not row or row[0] is None:
            continue
        class_label = str(row[0]).strip()
        matched_class = None
        for full, code in CONSTRUCTION_CLASSES:
            if class_label.startswith(full):
                matched_class = code
                break
        if matched_class is None:
            continue

        for col_idx, quality in quality_map.items():
            unit_cost = _safe_float(row[col_idx]) if col_idx < len(row) else None
            if unit_cost is None:
                continue
            pct_diff_col = pct_diff_map.get(col_idx)
            pct_diff = (
                _safe_float(row[pct_diff_col])
                if pct_diff_col is not None and pct_diff_col < len(row)
                else None
            )
            rows_out.append({
                "county_id":          county_id,
                "section_id":         section_id,
                "occupancy_code":     occupancy_code,
                "occupancy_desc":     occupancy_desc,
                "construction_class": matched_class,
                "quality_grade":      quality,
                "unit_cost":          unit_cost,
                "pct_diff_to_next":   pct_diff,
                "source_file":        file_name,
                "source_sheet":       sheet_name,
            })

    return rows_out


def seed_commercial(dry_run: bool = False) -> None:
    print(f"\n[2/6] costforge_commercial_schedules (Sections 11-18, 61, 64)")
    all_rows: list[dict] = []
    skipped: list[str] = []

    for section_id, file_name in COMMERCIAL_SECTIONS.items():
        path = SCHEDULES_DIR / file_name
        if not path.exists():
            skipped.append(file_name)
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            if sheet_name in ("Reference", "Template", "Test"):
                continue
            ws = wb[sheet_name]
            rows = _parse_commercial_sheet(
                ws, section_id, sheet_name, file_name, BENTON_COUNTY_ID
            )
            all_rows.extend(rows)
        wb.close()
        print(f"  Sec {section_id} ({file_name}): {len([r for r in all_rows if r['section_id']==section_id]):,} rows")

    if skipped:
        print(f"  Skipped missing files: {skipped}")
    bulk_upsert(
        "costforge_commercial_schedules", all_rows,
        "county_id,section_id,occupancy_code,construction_class,quality_grade",
        dry_run=dry_run
    )


# ── 3. Depreciation tables ────────────────────────────────────────────────────

def seed_depreciation(dry_run: bool = False) -> None:
    """
    Parse Depreciation.xlsx.
    Two sheets: 'Commercial Properties', 'Residential Properties'
    Row 1: header (prop type label)
    Row 2: matrix IDs (PACS cms_matrix IDs)
    Row 3: effective life in years (column headers: 70, 60, 55, 50, ...)
    Row 4+: age_years (col 0), then pct_good values per effective_life column
    """
    path = SCHEDULES_DIR / "Depreciation.xlsx"
    print(f"\n[3/6] costforge_depreciation ({path.name})")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows_out: list[dict] = []

    sheet_map = {
        "Commercial Properties": "commercial",
        "Residential Properties": "residential",
    }

    for sheet_name, prop_type in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: sheet '{sheet_name}' not found in {path.name}")
            continue
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))

        # Row 1 = description (skip)
        # Row 2 = matrix IDs → [col_idx: matrix_id]
        # Row 3 = effective life years → [col_idx: eff_life_yrs]
        # Rows 4+ = [age, pct_good × n_columns]

        if len(all_rows) < 3:
            continue

        matrix_id_row = all_rows[1]   # row index 1 (0-based)
        eff_life_row  = all_rows[2]   # row index 2

        # Build column map: col_idx → (pacs_matrix_id, effective_life_years)
        col_map: dict[int, tuple[int | None, int | None]] = {}
        for col_i in range(1, len(eff_life_row)):
            matrix_id = _safe_int(matrix_id_row[col_i]) if col_i < len(matrix_id_row) else None
            eff_life  = _safe_int(eff_life_row[col_i])
            if eff_life is not None:
                col_map[col_i] = (matrix_id, eff_life)

        for row in all_rows[3:]:
            if not row:
                continue
            age_years = _safe_int(row[0])
            if age_years is None:
                continue
            for col_i, (matrix_id, eff_life) in col_map.items():
                if col_i >= len(row):
                    continue
                pct_good = _safe_int(row[col_i])
                if pct_good is None:
                    continue
                rows_out.append({
                    "county_id":           BENTON_COUNTY_ID,
                    "property_type":       prop_type,
                    "age_years":           age_years,
                    "effective_life_years":eff_life,
                    "pct_good":            pct_good,
                    "pacs_matrix_id":      matrix_id,
                    "source_file":         path.name,
                })

    wb.close()
    bulk_upsert(
        "costforge_depreciation", rows_out,
        "county_id,property_type,age_years,effective_life_years",
        dry_run=dry_run
    )


# ── 4. Cost multipliers ───────────────────────────────────────────────────────

def seed_multipliers(dry_run: bool = False) -> None:
    """
    Parse local and current cost multipliers.xlsx.
    Sheet 1 (Local Multipliers sec99pg10):
      Col 0 = Class code, Col 1 = multiplier value
    Sheet 2 (Current Cost sec99pg3):
      Row 1: [None, 'Matrix ID', id1, id2, ...]
      Row 2: [None, None, 'Section 11', 'Section 12', ...]
      Row 3+: [None, class_description, mult1, mult2, ...]
    """
    path = SCHEDULES_DIR / "local and current cost multipliers.xlsx"
    print(f"\n[4/6] costforge_cost_multipliers ({path.name})")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows_out: list[dict] = []

    # Sheet 1: Local multipliers
    SECTION_CODE_MAP = {
        "Section 11": 11, "Section 12": 12, "Section 13": 13, "Section 14": 14,
        "Section 15": 15, "Section 16": 16, "Section 17": 17, "Section 18": 18,
    }
    CLASS_CODE_MAP = {"A": "A", "B": "B", "C": "C", "C ": "C", "D": "D", "S": "S", "P": "P"}

    local_sheet = wb.sheetnames[0]
    ws_local = wb[local_sheet]
    for row in ws_local.iter_rows(max_row=20, values_only=True):
        if not row or row[0] is None:
            continue
        class_raw = _clean_str(row[0])
        mult = _safe_float(row[1])
        if class_raw is None or mult is None:
            continue
        class_code = CLASS_CODE_MAP.get(class_raw)
        if class_code is None:
            continue
        rows_out.append({
            "county_id":         BENTON_COUNTY_ID,
            "multiplier_type":   "local",
            "construction_class":class_code,
            "section_id":        None,
            "multiplier":        mult,
            "source_file":       path.name,
        })

    # Sheet 2: Current cost multipliers
    if len(wb.sheetnames) >= 2:
        ws_curr = wb[wb.sheetnames[1]]
        all_rows = list(ws_curr.iter_rows(values_only=True))
        if len(all_rows) >= 3:
            matrix_id_row = all_rows[0]   # [None, 'Matrix ID', id1, id2, ...]
            section_name_row = all_rows[1] # [None, None, 'Section 11', ...]

            # Build col_idx → (section_id, pacs_matrix_id)
            curr_col_map: dict[int, tuple[int | None, int | None]] = {}
            for col_i in range(2, len(section_name_row)):
                sec_name = _clean_str(section_name_row[col_i])
                sec_id = SECTION_CODE_MAP.get(sec_name)
                matrix_id = _safe_int(matrix_id_row[col_i]) if col_i < len(matrix_id_row) else None
                if sec_id is not None:
                    curr_col_map[col_i] = (sec_id, matrix_id)

            for row in all_rows[2:]:
                if not row or row[1] is None:
                    continue
                class_desc = _clean_str(row[1])
                if class_desc is None:
                    continue
                # Map "Class A - Structural Steel Frame" → 'A'
                class_code = None
                for code in ["A", "B", "C", "D", "S", "P"]:
                    if f"Class {code}" in class_desc:
                        class_code = code
                        break
                if class_code is None:
                    continue
                for col_i, (sec_id, matrix_id) in curr_col_map.items():
                    mult = _safe_float(row[col_i]) if col_i < len(row) else None
                    if mult is None:
                        continue
                    rows_out.append({
                        "county_id":         BENTON_COUNTY_ID,
                        "multiplier_type":   "current",
                        "construction_class":class_code,
                        "section_id":        sec_id,
                        "multiplier":        mult,
                        "pacs_matrix_id":    matrix_id,
                        "source_file":       path.name,
                    })

    wb.close()
    bulk_upsert(
        "costforge_cost_multipliers", rows_out,
        "county_id,multiplier_type,construction_class,section_id",
        dry_run=dry_run
    )


# ── 5. Refinements (HVAC, Sprinklers) ────────────────────────────────────────

def seed_refinements(dry_run: bool = False) -> None:
    """
    Parse Refinement Matrices.xlsx — HVAC and Sprinklers sheets.
    HVAC sheet columns: Worksheet Ref, Proval Options, MVS Desc, PACS Naming,
                        Section 11, Section 12, ..., Section 18
    Each row = one HVAC type.
    Value = $/sqft add-on for that section.

    Sprinklers: section columns, area_band rows, value = $/sqft.
    """
    path = SCHEDULES_DIR / "Refinement Matrices.xlsx"
    print(f"\n[5/6] costforge_refinements ({path.name})")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows_out: list[dict] = []

    SECTION_COLS = {
        4: 11, 5: 12, 6: 13, 7: 14, 8: 15, 9: 16, 10: 17, 11: 18
    }

    # ── HVAC sheet ────────────────────────────────────────────────────────────
    if "HVAC" in wb.sheetnames:
        ws = wb["HVAC"]
        all_rows = list(ws.iter_rows(values_only=True))
        # Header row 0: Worksheet Ref, Proval Options, MVS Desc, PACS Naming, Sec11..18
        for row in all_rows[1:]:
            if not row or row[0] is None:
                continue
            worksheet_ref = _clean_str(row[0])
            pacs_naming   = _clean_str(row[3]) if len(row) > 3 else None
            qualifier     = pacs_naming or worksheet_ref
            if qualifier is None:
                continue
            for col_i, sec_id in SECTION_COLS.items():
                value = _safe_float(row[col_i]) if col_i < len(row) else None
                if value is None:
                    continue
                rows_out.append({
                    "county_id":      BENTON_COUNTY_ID,
                    "refinement_type":"hvac",
                    "qualifier":      qualifier,
                    "qualifier_desc": _clean_str(row[1]) if len(row) > 1 else None,
                    "section_id":     sec_id,
                    "area_band_min":  None,
                    "unit":           "per_sqft",
                    "value":          value,
                    "source_file":    path.name,
                    "source_sheet":   "HVAC",
                })

    # ── Sprinklers sheet ──────────────────────────────────────────────────────
    if "Sprinklers" in wb.sheetnames:
        ws = wb["Sprinklers"]
        all_rows = list(ws.iter_rows(values_only=True))
        # Row 0: headers — pairs of (None, Section X) for each section range
        # Row 1: Page reference (skip)
        # Data rows: col 0 = None, col 1 = area_band, then values in pattern
        # Section columns are paired: odd = section_id placeholder, next = value
        # Parse by finding which columns have section IDs
        if len(all_rows) >= 2:
            header_row = all_rows[0]
            # Map col_idx → section_id from header
            sprinkler_col_map: dict[int, int] = {}
            last_sec_id: int | None = None
            for col_i, cell in enumerate(header_row):
                sec_id = _safe_int(cell)
                if sec_id is not None and 11 <= sec_id <= 64:
                    last_sec_id = sec_id
                    sprinkler_col_map[col_i + 1] = sec_id  # value is in next column

            for row in all_rows[2:]:
                if not row:
                    continue
                area_band = _safe_int(row[1]) if len(row) > 1 else None
                for col_i, sec_id in sprinkler_col_map.items():
                    value = _safe_float(row[col_i]) if col_i < len(row) else None
                    if value is None:
                        continue
                    qualifier = f"sprinkler_sec{sec_id}"
                    rows_out.append({
                        "county_id":      BENTON_COUNTY_ID,
                        "refinement_type":"sprinkler",
                        "qualifier":      qualifier,
                        "qualifier_desc": f"Section {sec_id} sprinkler",
                        "section_id":     sec_id,
                        "area_band_min":  area_band,
                        "unit":           "per_sqft",
                        "value":          value,
                        "source_file":    path.name,
                        "source_sheet":   "Sprinklers",
                    })

    wb.close()
    bulk_upsert(
        "costforge_refinements", rows_out,
        "county_id,refinement_type,qualifier,section_id,area_band_min",
        dry_run=dry_run
    )


# ── 6. Improvement type codes ─────────────────────────────────────────────────

def seed_type_codes(dry_run: bool = False) -> None:
    """
    Parse PACS detail type codes-8-24.xlsx, sheet 'Main List'.
    Columns: imprv_det_type_cd, Change code to, imprv_det_typ_desc,
             Change Description to, Notes, Remove
    """
    path = SCHEDULES_DIR / "PACS detail type codes-8-24.xlsx"
    print(f"\n[6/6] costforge_imprv_type_codes ({path.name})")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Main List"] if "Main List" in wb.sheetnames else wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    rows_out: list[dict] = []
    # Header row
    for row in all_rows[1:]:
        if not row or row[0] is None:
            continue
        type_cd = _clean_str(row[0])
        if type_cd is None:
            continue
        canonical_code = _clean_str(row[1]) or type_cd
        type_desc      = _clean_str(row[2])
        canonical_desc = _clean_str(row[3]) or type_desc
        notes          = _clean_str(row[4])
        remove_flag    = _clean_str(row[5])

        # Determine if residential: residential types don't have section references
        # and typically reference dwelling components
        is_active = remove_flag is None or remove_flag.strip().upper() not in ("X", "X-")

        rows_out.append({
            "county_id":         BENTON_COUNTY_ID,
            "imprv_det_type_cd": type_cd,
            "type_desc":         type_desc,
            "canonical_code":    canonical_code,
            "canonical_desc":    canonical_desc,
            "is_active":         is_active,
            "notes":             notes,
        })

    bulk_upsert(
        "costforge_imprv_type_codes", rows_out,
        "county_id,imprv_det_type_cd",
        dry_run=dry_run
    )


# ── Entry point ───────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Seed Benton County CostForge schedules into TerraFusion Supabase")
    parser.add_argument("--dry-run", action="store_true",
                        help="Parse & print counts — do NOT write to Supabase")
    parser.add_argument("--table", type=str, default=None,
                        choices=["residential", "commercial", "depreciation",
                                 "multipliers", "refinements", "type_codes"],
                        help="Run only a single table seeder")
    args = parser.parse_args()

    if not SERVICE_KEY and not args.dry_run:
        print("[seed_costforge] ERROR: SUPABASE_SERVICE_ROLE_KEY not set.")
        sys.exit(1)

    if not SCHEDULES_DIR.exists():
        print(f"[seed_costforge] ERROR: Schedules dir not found: {SCHEDULES_DIR}")
        sys.exit(1)

    t_start = time.time()

    run_all = args.table is None
    if run_all or args.table == "residential":
        seed_residential(dry_run=args.dry_run)
    if run_all or args.table == "commercial":
        seed_commercial(dry_run=args.dry_run)
    if run_all or args.table == "depreciation":
        seed_depreciation(dry_run=args.dry_run)
    if run_all or args.table == "multipliers":
        seed_multipliers(dry_run=args.dry_run)
    if run_all or args.table == "refinements":
        seed_refinements(dry_run=args.dry_run)
    if run_all or args.table == "type_codes":
        seed_type_codes(dry_run=args.dry_run)

    print(f"\n✓ CostForge seeder complete in {time.time()-t_start:.1f}s")


if __name__ == "__main__":
    main()
